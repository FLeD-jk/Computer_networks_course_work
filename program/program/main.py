import pickle
import random as rnd
from tkinter import *
from tkinter.tix import Tk, HList, Scrollbar
import tkinter.tix as tk
from routing_algorithm_catalogs import *
from channel import Channel
from graph import *
from node import Node

from workstation import Station
from tkinter import Toplevel, Label, Radiobutton, IntVar, Button, Entry
import openpyxl
import os
from openpyxl import load_workbook, Workbook

global CURRENT_NODE_NUMBER
global CHANNEL_WEIGHTS

class Package:
    def __init__(self, size, number, header=24):
        self.number = number
        self.size = size
        self.header = header

    def __repr__(self):
        return f"Package(number={self.number} size={self.size} header={self.header})"

class Main:
    def __init__(self):

        self.node_list = []
        self.channel_list = []
        # left panel
        self.root = Tk()
        self.root["bg"] = "#ffffff"
        self.root.title("Course Work Nesteruk Anastasia KV-11")
        self.root.geometry("1024x520")

        # background

        self.left_panel = Frame(self.root, width=128, height=2000, bg="#e1dd72")
        self.left_panel.pack(side="right", anchor="nw")

        self.canvas = Canvas(width=1280, height=1024, bg='#DCDCDC')
        self.canvas.pack(side="right", anchor="nw")

        # adding buttons
        self.button_add_node = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", text="Add Node", font=("Time New Roman", 9))
        self.button_add_node.place(x=902, y=10)
        self.button_add_node.bind('<Button-1>', self.add_node_event)

        self.button_add_node = Button(self.root, width=15, height=2, bg="#a8c66c",  relief="ridge", fg="black", text="Add workstation", font=("Time New Roman", 9))
        self.button_add_node.place(x=902, y=60)
        self.button_add_node.bind('<Button-1>', self.add_station_event)

        self.button_add_channel = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", text="Add channel", font=("Time New Roman", 9))
        self.button_add_channel.place(x=902, y=110)
        self.button_add_channel.bind('<Button-1>', self.add_channel)

        self.button_delete_node = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", text="Delete node", font=("Time New Roman", 9))
        self.button_delete_node.place(x=902, y=160)
        self.button_delete_node.bind('<Button-1>', self.delete_node)

        self.button_delete_node = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", bd=2,  text="Delete workstation", font=("Time New Roman", 9))
        self.button_delete_node.place(x=902, y=210)
        self.button_delete_node.bind('<Button-1>', self.delete_node)

        self.button_delete_channel = Button(self.root, width=15, height=2, bg="#a8c66c",relief="ridge", fg="black", bd=2, text="Delete channel", font=("Time New Roman", 9))
        self.button_delete_channel.place(x=902, y=260)
        self.button_delete_channel.bind('<Button-1>', self.delete_channel)

        self.delete_everything_button = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", bd=2, text="Delete everything", font=("Time New Roman", 9))
        self.delete_everything_button.place(x=902, y=310)
        self.delete_everything_button.bind('<Button-1>', self.clear_everything)

        self.button_send_message = Button(self.root, width=15, height=2, wraplength=100, bg="#a8c66c", relief="ridge", fg="black", bd=2, text="Send message", font=("Time New Roman", 9))
        self.button_send_message.place(x=902, y=360)
        self.button_send_message.bind('<Button-1>', self.send_message)
        '''
        self.save_to_file_button = Button(self.root, width=15, height=2, bg="#a8c66c", relief="ridge", fg="black", bd=2, text="Save", font=("Time New Roman", 9))
        self.save_to_file_button.place(x=902, y=846)
        self.save_to_file_button.bind('<Button-1>', self.save_to_file)

        self.upload_from_file_button = Button(self.root, width=15, height=3, bg="#a8c66c", relief="ridge", fg="black", bd=2, text="Load half-duplex \nnetwork", font=("Time New Roman", 9))
        self.upload_from_file_button.place(x=902, y=896)
        self.upload_from_file_button.bind('<Button-1>', self.upload_from_file)

        self.upload_from_file_button1 = Button(self.root, width=15, height=3, bg="#a8c66c",relief="ridge", fg="black", bd=2,
                                               text="Load duplex \nnetwork", font=("Time New Roman", 9))
        self.upload_from_file_button1.place(x=902, y=960)
        self.upload_from_file_button1.bind('<Button-1>', self.upload_from_file1)
        '''
        self.title = ''
        self.root.mainloop()

    def save_to_file(self, event):
        self.clear()
        with open("nodes.dat", "wb") as file_nodes:
            for node_ in self.node_list:
                pickle.dump([node_.x, node_.y, node_.name, node_.type], file_nodes)
        with open("channels.dat", "wb") as file_channels:
            for channel_ in self.channel_list:
                if channel_.is_satellite:
                    pickle.dump([channel_.first_node, channel_.second_node, channel_.type, channel_.x1, channel_.y1, channel_.x2,
                                 channel_.y2, channel_.weight/3, channel_.is_satellite], file_channels)
                else:
                    pickle.dump([channel_.first_node, channel_.second_node, channel_.type, channel_.x1, channel_.y1, channel_.x2, channel_.y2, channel_.weight, channel_.is_satellite], file_channels)


    def upload_from_file(self, event):
        self.clear()
        global CURRENT_NODE_NUMBER
        CURRENT_NODE_NUMBER = 0
        with open("nodes.dat", "rb") as file1:
            try:
                while True:
                    node_ = pickle.load(file1)
                    if node_[3] == "node":
                        new_view_node = Node(self.canvas, node_[2], node_[0], node_[1])
                    else:
                        new_view_node = Station(self.canvas, node_[2], node_[0], node_[1])
                    self.node_list.append(new_view_node)
                    CURRENT_NODE_NUMBER += 1
            except EOFError:
                pass
        with open("channels.dat", "rb") as file_channels:
            try:
                while True:
                    channel_ = pickle.load(file_channels)
                    new_channel = Channel(self.canvas, [channel_[0], channel_[1]],
                                          channel_[2], [channel_[3], channel_[4],
                                                              channel_[5], channel_[6]],
                                          int(channel_[7]), channel_[8])
                    for element in self.node_list:
                        if element.name == channel_[0] or element.name == channel_[1]:
                            element.related_channels.append(new_channel)
                    self.channel_list.append(new_channel)

                    for element in self.node_list:
                        if element.selected:
                            element.selected = False
                            self.canvas.itemconfig(element.view, fill=element.color)
            except EOFError as ex:
                print(ex)

    def upload_from_file1(self, event):
        self.clear()
        global CURRENT_NODE_NUMBER
        CURRENT_NODE_NUMBER = 0
        with open("nodes1.dat", "rb") as file1:
            try:
                while True:
                    node_ = pickle.load(file1)
                    if node_[3] == "node":
                        new_view_node = Node(self.canvas, node_[2], node_[0], node_[1])
                    else:
                        new_view_node = Station(self.canvas, node_[2], node_[0], node_[1])
                    self.node_list.append(new_view_node)
                    CURRENT_NODE_NUMBER += 1
            except EOFError:
                pass
        with open("channels1.dat", "rb") as file_channels:
            try:
                while True:
                    channel_ = pickle.load(file_channels)
                    new_channel = Channel(self.canvas, [channel_[0], channel_[1]],
                                          channel_[2], [channel_[3], channel_[4],
                                                              channel_[5], channel_[6]],
                                          int(channel_[7]), channel_[8])
                    for element in self.node_list:
                        if element.name == channel_[0] or element.name == channel_[1]:
                            element.related_channels.append(new_channel)
                    self.channel_list.append(new_channel)

                    for element in self.node_list:
                        if element.selected:
                            element.selected = False
                            self.canvas.itemconfig(element.view, fill=element.color)
            except EOFError as ex:
                print(ex)

    def add_node_event(self, event):
        self.canvas.unbind('<Button-1>')
        self.clear()
        self.canvas.bind('<Button-1>', lambda e: None)
        self.add_node(event)

    def add_node(self, event):
        global CURRENT_NODE_NUMBER
        new_view_node = Node(self.canvas, CURRENT_NODE_NUMBER, event.x, event.y)
        self.node_list.append(new_view_node)
        CURRENT_NODE_NUMBER += 1

    def add_station_event(self, event):
        self.canvas.unbind('<Button-1>')
        self.clear()
        self.canvas.bind('<Button-1>', lambda e: None)
        self.add_station(event)


    def add_station(self, event):
        global CURRENT_NODE_NUMBER
        new_view_node = Station(self.canvas, CURRENT_NODE_NUMBER, event.x, event.y)
        self.node_list.append(new_view_node)
        CURRENT_NODE_NUMBER += 1

    def destoy_channel_text(self):
        self.choose_label_2.destroy()
        self.confirm_button.destroy()
        self.random.destroy()
        self.manually.destroy()
        self.duplex.destroy()
        self.satellite.destroy()
        self.ground.destroy()
        self.half_duplex.destroy()

    def add_channel(self, event):
        self.clear()
        self.nodes_for_channel = [node for node in self.node_list if node.selected]

        if len(self.nodes_for_channel) == 2:
            self.channel_window = Toplevel(self.root)
            self.channel_window.geometry("300x500")
            self.channel_window.title("Channel Settings")
            self.channel_window["bg"] = "#e1dd72"

            # Select the channel type (satellite or regular)
            Label(self.channel_window, text="Chose channel mode:", font=("Time New Roman", 14), bg="#e1dd72").pack(pady=10)
            self.is_satellite = IntVar(value=0)
            Radiobutton(self.channel_window, text="Simple channel", font=("Time New Roman", 9),bg="#e1dd72", variable=self.is_satellite, value=0).pack(anchor=W,
                                                                                                              padx=20)
            Radiobutton(self.channel_window, text="Satellite channel", font=("Time New Roman", 9), bg="#e1dd72", variable=self.is_satellite, value=1).pack(
                anchor=W, padx=20)

            # Selecting the channel type (duplex or half-duplex)
            Label(self.channel_window, text="Chose channel type:",  font=("Time New Roman", 14),bg="#e1dd72").pack(pady=10)
            self.channel_type = IntVar(value=0)
            Radiobutton(self.channel_window, text="Duplex", font=("Time New Roman", 9),bg="#e1dd72", variable=self.channel_type, value=0).pack(anchor=W, padx=20)
            Radiobutton(self.channel_window, text="Half-Duplex", font=("Time New Roman", 9), bg="#e1dd72", variable=self.channel_type, value=1).pack(anchor=W,
                                                                                                           padx=20)

            # Selecting a weight setting method
            Label(self.channel_window, text="Chose weight setting:",  font=("Time New Roman", 14), bg="#e1dd72").pack(pady=10)
            self.weight_type = IntVar(value=0)
            Radiobutton(self.channel_window, text="Random",font=("Time New Roman", 9),bg="#e1dd72", variable=self.weight_type, value=0).pack(anchor=W, padx=20)
            Radiobutton(self.channel_window, text="Manually", font=("Time New Roman", 9), bg="#e1dd72", variable=self.weight_type, value=1).pack(anchor=W,
                                                                                                       padx=20)

            # Confirmation button
            Button(self.channel_window, text="Confirm Settings",font=("Time New Roman", 9), bg="#a8c66c", command=self.confirm_channel).pack(pady=20)

    def confirm_channel(self):
        channel_mode = "duplex" if self.channel_type.get() == 0 else "half-duplex"
        is_satellite_mode = bool(self.is_satellite.get())
        weight_mode = self.weight_type.get()

        if weight_mode == 0:
            weight = rnd.choice(CHANNEL_WEIGHTS)
            self.create_new_channel(channel_mode, weight, is_satellite_mode)
            self.channel_window.destroy()
        else:
            self.prompt_manual_weight(channel_mode, is_satellite_mode)

    def prompt_manual_weight(self, channel_mode, is_satellite_mode):
        # Delete the old input field (if any)
        for widget in self.channel_window.winfo_children():
            if isinstance(widget, Entry):
                widget.destroy()

        # Create a new field for entering weight
        manual_weight_label = Label(self.channel_window, text="Enter weight:",  font=("Time New Roman", 14), bg="#e1dd72")
        manual_weight_label.pack(pady=5)

        self.manual_weight_entry = Entry(self.channel_window, bg='#DCDCDC')
        self.manual_weight_entry.pack(pady=5)

        # Button to confirm the manually entered weight
        self.confirm_weight_button = Button(self.channel_window, text="Confirm weight",  font=("Time New Roman", 9),  bg="#a8c66c",
                                            command=lambda: self.confirm_manual_weight(channel_mode, is_satellite_mode))
        self.confirm_weight_button.pack(pady=10)

    def confirm_manual_weight(self, channel_mode, is_satellite_mode):
        try:
            weight = int(self.manual_weight_entry.get())
            self.create_new_channel(channel_mode, weight, is_satellite_mode)
            self.channel_window.destroy()
        except ValueError:
            error_label = Label(self.channel_window, text="Invalid weight. Enter an integer.", fg="red",  bg="#e1dd72",
                                font=("Time New Roman", 14))
            error_label.pack()

    def create_new_channel(self, channel_type, weight, is_satellite):
        new_channel = Channel(
            self.canvas,
            [self.nodes_for_channel[0].name, self.nodes_for_channel[1].name],
            channel_type,
            [self.nodes_for_channel[0].x, self.nodes_for_channel[0].y, self.nodes_for_channel[1].x,
             self.nodes_for_channel[1].y],
            weight,
            is_satellite=is_satellite
        )
        self.channel_list.append(new_channel)
        for node in self.nodes_for_channel:
            node.related_channels.append(new_channel)
            node.selected = False
            self.canvas.itemconfig(node.view, fill=node.color)

    def delete_channel(self, event):
        self.clear()
        for channel in self.channel_list:
            if channel.selected:
                channel.delete()
                self.channel_list.remove(channel)
                break
        for element in self.node_list:
            if channel in element.related_channels:
                element.related_channels.remove(channel)

    def delete_node(self, event):
        self.clear()
        channels_for_delete = []
        for element in self.node_list:
            if element.selected:
                element.delete()
                for channel in self.channel_list:
                    if channel in element.related_channels:
                        element.related_channels.remove(channel)
                        channels_for_delete.append(channel)
                        channel.delete()
                self.node_list.remove(element)
            for channel in channels_for_delete:
                self.channel_list.remove(channel)

    def clear_everything(self, event):
        for channel in self.channel_list:
            channel.delete()
        for node in self.node_list:
            node.delete()
        self.channel_list.clear()
        self.node_list.clear()

    def clear(self):
        for channel in self.channel_list:
            channel.canvas.itemconfig(channel.view, fill=channel.current_color, width=channel.width)

    def send_message(self, event):
        self.clear()
        self.nodes_for_message = [node for node in self.node_list if node.selected]

        if len(self.nodes_for_message) == 1:
            # Create a new window for entering message parameters
            self.message_window = Toplevel(self.root)
            self.message_window.geometry("300x400")
            self.message_window.title("Send Message")
            self.message_window["bg"] = "#e1dd72"

            # Enter the size of the message
            self.size_label = Label(self.message_window, text="Message size (bytes):", font=("Time New Roman", 14),  bg="#e1dd72")
            self.size_label.pack(pady=10)
            self.size_enter = Entry(self.message_window, width=10, bd=3, bg='#DCDCDC')
            self.size_enter.pack(pady=10)

            # Entering the packet size
            self.package_size_label = Label(self.message_window, text="Info package size (bytes):", font=("Time New Roman", 14), bg="#e1dd72")
            self.package_size_label.pack(pady=10)
            self.package_size_enter = Entry(self.message_window, width=10, bd=3, bg='#DCDCDC')
            self.package_size_enter.pack(pady=10)

            # Select the type of channel for transmission
            self.send_type = IntVar()
            self.send_type.set(0)
            self.datagram = Radiobutton(self.message_window, text="Datagram",font=("Time New Roman", 9), bg="#e1dd72", variable=self.send_type, value=0)
            self.virtual = Radiobutton(self.message_window, text="Virtual channel", font=("Time New Roman", 9), bg="#e1dd72", variable=self.send_type, value=2)
            self.datagram.pack(pady=5)
            self.virtual.pack(pady=5)

            # Confirmation button
            self.confirm_button = Button(self.message_window,  text="OK",font=("Time New Roman", 9),  bg="#a8c66c",
                                         command=lambda event=None: self.confirm_send(event))
            self.confirm_button.pack(pady=20)


    def create_graph(self, graph):
        for channel in self.channel_list:
            if not channel.disabled:
                flag = True
                for node in self.node_list:
                    if (node.name == channel.second_node and node.disabled) or (node.name == channel.first_node and node.disabled):
                        flag = False
                        break
                if flag:
                    graph.add_edge(channel.first_node, channel.second_node, channel.weight)
        return graph.edges
    

    def transform_route(self, route):
        i = 0
        route_channels = []
        for element in route:
            if i == 0:
                pass
            else:
                for channel in self.channel_list:
                    if (channel.first_node == route[i - 1] and channel.second_node == route[i]) or (channel.second_node == route[i - 1] and channel.first_node == route[i]):
                        route_channels.append(channel)
                        break
            i += 1
        return route_channels

    def confirm_send(self, event):
        self.message_size = self.size_enter.get()
        self.package_size = self.package_size_enter.get()
        if self.message_size == "" or self.package_size == "":
            error_label = Label(self.message_window, text="You need to enter all fields", fg="red",  bg="#e1dd72",
                                font=("Time New Roman", 14))
            error_label.pack()
        else:
            self.message_size = int(self.message_size, 10)
            print(self.message_size)
            self.package_size = int(self.package_size, 10)
            if self.message_size < 1 or self.message_size > 65532:
                error_label = Label(self.message_window, text="Wrong message size", fg="red",  bg="#e1dd72",
                                    font=("Time New Roman", 14))
                error_label.pack()
            elif self.package_size < 140: #or self.package_size > self.message_size:
                error_label = Label(self.message_window, text="Wrong package size", fg="red", bg="#e1dd72",
                                    font=("Time New Roman", 14))
                error_label.pack()
            else:
                self.confirm_button.destroy()
                self.datagram.destroy()
                # self.connection.destroy()
                self.virtual.destroy()
                self.size_label.destroy()
                self.package_size_label.destroy()
                self.size_enter.destroy()
                self.package_size_enter.destroy()
                for element in self.node_list:
                    if element.selected:
                        element.selected = False
                        self.canvas.itemconfig(element.view, fill=element.color)


                # add size of package
                # header size depends on type of channel
                # case simple

                if self.send_type.get() == 0:
                    self.title = "Datagram results"
                    self.header_size = 28
                    # case hal-duplex
                ##elif self.send_type.get() == 1:
                ##    self.title = "Logic channel results"
                ##    self.header_size = 40
                    # case virtual
                elif self.send_type.get() == 2:
                    self.title = "Virtual channel result"
                    self.header_size = 5

                # informational data
                self.data_package = self.package_size - self.header_size

                # quantity of info packages
                self.info_package_quantity = 0
                if self.message_size % self.data_package == 0:
                    self.info_package_quantity = self.message_size // self.data_package
                else:
                    self.info_package_quantity = self.message_size // self.data_package + 1

                self.package_list = []
                
                i = 1
                while i <= self.info_package_quantity:
                    if i != self.info_package_quantity:
                        self.package_list.append(Package(self.data_package, i, self.header_size))
                    else:
                        if self.message_size % self.data_package != 0:
                            self.package_list.append(Package(self.message_size % self.data_package, i, self.header_size))
                        else:
                            self.package_list.append(Package(self.data_package, i, self.header_size))
                    i += 1

                print(self.package_list)
                
                app = Tk()
                app["bg"] = "#e1dd72"
                app.title(self.title)
                app.minsize(1100, 180)
                app.maxsize()
                app.rowconfigure(0, weight=1)
                app.columnconfigure(0, weight=1)

                tab = HList(app, columns=3, header=True, width=150)
                tab.grid(row=0, column=0, sticky="nswe")
                tab.grid(row=0, column=1, sticky="nswe")
                scroll = Scrollbar(app, command=tab.yview)
                tab['yscrollcommand'] = scroll.set
                scroll.grid(row=0, column=1, sticky="nwse")
                tab["bg"] = '#DCDCDC'
                # creating headers
                tab.header_create(0, text="Destination")
                tab.header_create(1, text="Time")
                tab.header_create(2, text="Path")
                self.total_time_to_all_nodes = 0

                counter = 0
                tab_marsh = defaultdict(list)

                # 2.1 move variables above cycle
                self.service_package_count = 0
                self.quantity_info_package = 0
                self.service_size_info = 0
                self.all_trafic = 0

                for finish_node in self.node_list:

                    self.graph = Graph()
                    self.graph_edges = self.create_graph(self.graph)
                    # 2.1 add skip node if this not station
                    if self.nodes_for_message[0].name == finish_node.name or finish_node.type != "station":
                        continue
                    print("From ", self.nodes_for_message[0].name, " node ", " to ", finish_node.name, " node")
                    
                    self.route = session_catalog_routing(self.graph.edges, self.nodes_for_message[0].name, finish_node.name)
                    counter += 1

                    self.route_channels = self.transform_route(self.route) 

                    # generation error in channel of route
                    for channel in self.route_channels:
                        if rnd.random() <= channel.error_prob:
                            print ("Error in graph:" , channel.first_node, channel.second_node)
                            
                            # deleting from graph channels with error
                            for edge in self.graph.edges:
                                if (edge[0]==channel.first_node and edge[1]==channel.second_node):
                                    self.graph.edges.remove(edge)
                                    break
                            for edge in self.graph.edges:
                                if (edge[0]==channel.second_node and edge[1]==channel.first_node):
                                    self.graph.edges.remove(edge)
                                    break

                            # search for new node from which we create route
                            i=0
                            for element in self.route:
                                if i==0:
                                    pass
                                else:
                                    if (channel.first_node == self.route[i - 1] and channel.second_node == self.route[i]):
                                        self.new_start_node = channel.first_node
                                        break
                                    elif(channel.second_node == self.route[i - 1] and channel.first_node == self.route[i]):
                                        self.new_start_node = channel.second_node
                                        break
                                i+=1
                                
                            #deleting error route
                            self.old_route = self.route
                            self.route = []
                            for element in self.old_route:
                                if element != self.new_start_node:
                                    self.route.append(element)
                                else:
                                    break

                            # building new part of route and adding it to main route instead of old one
                            self.new_route_part = session_catalog_routing(self.graph.edges, self.new_start_node, finish_node.name)
                            for element in self.new_route_part:
                                self.route.append(element)
                            self.route_channels = self.transform_route(self.route) 

                    print("Result route: ", " --> ".join([str(x) for x in self.route]))
                    marsh = " --> ".join([str(x) for x in self.route])
                    tab_marsh[self.nodes_for_message[0].name].append(marsh)                

                    self.time = 0
                    
                    # in case of data 
                    if self.send_type.get() == 0:
                        for channel in self.route_channels:
                            if channel.type == "half-duplex":
                                # 2.1 add random to half duplex
                                slow = round(rnd.randint(190, 210)/100, 2)
                            else:
                                slow = 1
                            i = 1
                            for package in self.package_list:
                                self.time += 2 * (package.size + package.header) // 100 * channel.weight * slow
                                if (package.size + package.header) % 100 != 0:
                                    self.time += 2* ((package.size + package.header) % 100) / 100 * channel.weight * slow

                                x2_header = 2 * package.header

                                self.service_size_info += x2_header
                                self.all_trafic += package.size + x2_header
                                self.service_package_count += 1
                                self.quantity_info_package += 1 
                        

                    elif self.send_type.get() == 1:
                        for channel in self.route_channels:
                            if channel.type == "half-duplex":
                                # 2.1 add random to half duplex
                                slow = round(rnd.randint(190, 210)/100, 2)
                            else:
                                slow = 1
                            
                            # 2.1  add service package into package list
                            service_packages = [Package(0, x, 40) for x in range(4)]

                            # 2.1 add merge   
                            for package in [*self.package_list, *service_packages] :

                                self.time += 2 *(package.size + package.header) // 100 * channel.weight * slow
                                if (package.size + package.header) % 100 != 0:
                                    self.time += 2 *((package.size + package.header) % 100) / 100 * channel.weight * slow
                                
                                # 2.1 add counts below
                                x2_header = 2 * package.header

                                self.service_size_info += x2_header
                                self.all_trafic += package.size + x2_header
                                self.service_package_count += 1
                                self.quantity_info_package += 1

                            # 2.1 sub package on set connection
                            self.quantity_info_package -= 4 
                            self.service_package_count += 4

                    elif self.send_type.get() == 2:
                        
                        for channel in self.route_channels:
                            if channel.type == "half-duplex":
                                # 2.1 add random to half duplexx
                                slow = round(rnd.randint(190, 210)/100, 2)
                            else:
                                slow = 1

                            # 2.1  add service package into package list    
                            service_packages = [Package(0, x, 40) for x in range(4)]

                            # 2.1 ADD MERGE    
                            for package in [*self.package_list, *service_packages]:

                                self.time += 2 * (package.size + package.header) // 100 * channel.weight * slow * 1.37
                                if (package.size + package.header) % 100 != 0:
                                    self.time += 2 * ((package.size + package.header) % 100) / 100 * channel.weight * slow * 1.37
                                
                        
                                x2_header = 2 * package.header

                                self.service_size_info += x2_header
                                self.all_trafic += package.size + x2_header
                                self.service_package_count += 1
                                self.quantity_info_package += 1 

                            # 2.1 sub package on set connection
                            self.quantity_info_package -= 4                             
                            self.service_package_count += 4

                    if finish_node.type == "station":
                        index = '%s' % counter
                        tab.add(index, data="--<%s>--" % counter)


                        time_to_node = round(self.time / 1000, 2)
                        self.total_time_to_all_nodes += time_to_node

                        tab.item_create(index, 0, text=(finish_node.name))
                        if self.route != []:
                            tab.item_create(index, 1, text=(f"{time_to_node} sec."))
                            tab.item_create(index, 2, text=marsh)
                        else:
                            tab.item_create(index, 1, text=("Cannot find way."))
                            tab.item_create(index, 2, text=marsh)

                    self.route_channels = []
                    self.route = []

                pack_quantity_lab1 = Label(app, text="Number of information packets: ",font=("Time New Roman", 12),  bg="#e1dd72")
                pack_quantity_lab1.place(x=5, y=30)
                pack_quantity_lab2 = Label(app, text=str(self.quantity_info_package), font=("Time New Roman", 12),  bg="#e1dd72")
                pack_quantity_lab2.place(x=270, y=30)

                servise_data_lab1 = Label(app, text="Number of service packages: ", font=("Time New Roman", 12),  bg="#e1dd72")
                servise_data_lab1.place(x=5, y=0)
                servise_data_lab2 = Label(app, text=str(self.service_package_count), font=("Time New Roman", 12),  bg="#e1dd72")
                servise_data_lab2.place(x=270, y=0)


                from_node_lab1 = Label(app, text="Volume of information traffic: ", font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab1.place(x=5, y=60)
                from_node_lab2 = Label(app, text=str(self.all_trafic - self.service_size_info) + " bytes",font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab2.place(x=270, y=60)

                # 2.1 add service data
                from_node_lab1 = Label(app, text="Volume of service traffic: ",font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab1.place(x=5, y=90)
                from_node_lab2 = Label(app, text=str(self.service_size_info) + " bytes", font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab2.place(x=270, y=90)

                
                from_node_lab1 = Label(app, text="General traffic: ", font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab1.place(x=5, y=120)
                from_node_lab2 = Label(app, text=str(self.all_trafic) + " bytes", font=("Time New Roman", 12),  bg="#e1dd72")
                from_node_lab2.place(x=270, y=120)

                total_time_1 = Label(app, text="Time: ", font=("Time New Roman", 12),  bg="#e1dd72")
                total_time_1.place(x=5, y=150)
                total_time_2 = Label(app, text=str(int(self.total_time_to_all_nodes)) + " seconds.", font=("Time New Roman", 12),  bg="#e1dd72")
                total_time_2.place(x=270, y=150)
                        
                tab.place(x=370, y=0)
                print(tab_marsh)


                results = []
                transmission_mode = ''
                if self.send_type.get() == 2:
                    transmission_mode = "Virtual channel"
                if self.send_type.get() == 0:
                    transmission_mode = "Datagram"

                results.append({
                    "transmission_mode": transmission_mode,
                    "message_size": self.message_size,
                    "info_packets_size": self.package_size,
                    "info_packets": self.quantity_info_package,
                    "service_packets": self.service_package_count,
                    "info_traffic": self.all_trafic - self.service_size_info,
                    "service_traffic": self.service_size_info,
                    "general_traffic": self.all_trafic,
                    "time": self.total_time_to_all_nodes
                })

                self.save_results_to_xlsx("network_simulation_results.xlsx", results)

    def save_results_to_xlsx(self, file_name, results):
        """
        Save results to an XLSX file, appending new results to the end of the file.

        Parameters:
        - file_name: str, name of the output XLSX file.
        - results: list of dicts, where each dict contains results for one message.
        """
        if os.path.exists(file_name):  # Check if file exists
            wb = load_workbook(file_name)  # Load existing workbook
            ws = wb.active
        else:
            wb = Workbook()  # Create a new workbook if file does not exist
            ws = wb.active
            ws.title = "Results"

            # Write the header only if the file is new
            headers = [
                "Тип",
                "Розмір повідомлення",
                "Розмір інформаційного пакету",
                "Кількість інформаційних пакетів",
                "Кількість службових пакетів",
                "Розмір інформаційного пакету,байт",
                "Розмір службового трафіку (байт)",
                "Загальний трафік",
                "Час"
            ]
            ws.append(headers)

        # Write the results to the next available row
        for result in results:
            ws.append([
                result.get("transmission_mode"),
                result.get("message_size"),
                result.get("info_packets_size"),
                result.get("info_packets"),
                result.get("service_packets"),
                result.get("info_traffic"),
                result.get("service_traffic"),
                result.get("general_traffic"),
                result.get("time")
            ])

        # Save the workbook with the new results
        wb.save(file_name)



if __name__ == '__main__':
    CHANNEL_WEIGHTS = [1, 2, 3, 5, 7, 11, 12, 15, 18, 32]
    CURRENT_NODE_NUMBER = 0
    ITERATION = 1
    Main()
